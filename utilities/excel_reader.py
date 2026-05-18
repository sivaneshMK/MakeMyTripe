from pathlib import Path

from openpyxl.reader.excel import load_workbook


class ExcelReader:

    @staticmethod
    def get_test_data(file_path, sheet_name, test_name):


        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        header = []
        for cell in sheet[1]:
            header.append(cell.value)

        test_data = {}

        for row in sheet.iter_rows(min_row=2, values_only=True):

            row_data = dict(zip(header, row))

            if row_data["Test Name"] == test_name:
                test_data = row_data
                break
        wb.close()
        return test_data

    @staticmethod
    def get_multiple_test_data(sheet_name, test_name):

        root_path = Path(__file__).parent.parent
        file_path = root_path / "test_data" / "test_data.xlsx"
        wb = load_workbook(file_path)
        sheet = wb[sheet_name]

        header = []
        for cell in sheet[1]:
            header.append(cell.value)

        test_data = []

        for row in sheet.iter_rows(min_row=2, values_only=True):

            row_data = dict(zip(header, row))

            if row_data["Test Name"] == test_name:
                test_data.append(row_data)


        wb.close()
        print(test_data)
        return test_data